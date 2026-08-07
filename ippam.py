import datetime
import hashlib
import io
import ipaddress
import json
import os
import socket
import sqlite3
import subprocess
import urllib.request
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

# ==========================================
# 1. CONFIGURAÇÃO INICIAL E ESTILIZAÇÃO CSS
# ==========================================
st.set_page_config(
    page_title="IPAM Enterprise - Gestão, Automação e Engenharia NOC",
    layout="wide",
    page_icon="🌐",
)

st.markdown("""
<style>
    .stApp {
        background-color: #0d1117;
        color: #c9d1d9;
    }
    div[data-testid="stMetric"] {
        background-color: #161b22;
        border: 1px solid #30363d;
        border-radius: 8px;
        padding: 15px 20px;
        box-shadow: 0 4px 10px rgba(0, 0, 0, 0.4);
        transition: transform 0.2s ease, border-color 0.2s ease;
    }
    div[data-testid="stMetric"]:hover {
        transform: translateY(-3px);
        border-color: #58a6ff;
    }
    div[data-testid="stMetricLabel"] {
        color: #8b949e !important;
        font-weight: 600;
        font-size: 0.9rem;
    }
    div[data-testid="stMetricValue"] {
        color: #58a6ff !important;
        font-weight: bold;
    }
    .stButton>button {
        border-radius: 6px;
        font-weight: 600;
        transition: all 0.2s ease;
    }
    .stCodeBlock {
        border-left: 4px solid #238636 !important;
        border-radius: 4px;
    }
    div[data-testid="stDataFrame"] {
        border: 1px solid #30363d;
        border-radius: 8px;
        background-color: #161b22;
    }
</style>
""", unsafe_allow_html=True)

components.html("""
<script>
const doc = window.parent.document;
doc.addEventListener('keydown', function(e) {
    if ((e.ctrlKey || e.metaKey) && e.key === 'k') {
        e.preventDefault();
        const input = doc.querySelector('input[aria-label*="Digite um IP"]');
        if (input) input.focus();
    }
});
</script>
""", height=0)


def botao_copiar(texto_para_copiar, id_unico="btnCopy"):
  js_code = f"""
    <script>
    function copyText_{id_unico}() {{
        navigator.clipboard.writeText({json.dumps(texto_para_copiar)});
        let btn = document.getElementById('{id_unico}');
        btn.innerText = "✅ Copiado!";
        btn.style.backgroundColor = "#2ea043";
        setTimeout(() => {{ 
            btn.innerText = "📋 Copiar Configuração"; 
            btn.style.backgroundColor = "#238636";
        }}, 2000);
    }}
    </script>
    <button id="{id_unico}" onclick="copyText_{id_unico}()" style="
        background-color: #238636;
        color: #ffffff;
        border: 1px solid rgba(240,246,252,0.1);
        padding: 8px 16px;
        border-radius: 6px;
        cursor: pointer;
        font-weight: 600;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
        margin-top: 5px;
        margin-bottom: 10px;">
        📋 Copiar Configuração
    </button>
    """
  components.html(js_code, height=50)


DB_FILE = "ipam_database.db"

COLUNAS_ASN = ["ASN", "Entidade", "Tipo", "Contato"]
COLUNAS_IP = [
    "Prefixo",
    "Versão",
    "ASN Vinculado",
    "Finalidade",
    "VLAN ID",
    "VRF",
    "POP / Localidade",
    "ID Circuito",
]


# ==========================================
# 2. SEGURANÇA E HASH DE SENHA
# ==========================================
def make_hashes(password):
  return hashlib.sha256(str.encode(password)).hexdigest()


def check_hashes(password, hashed_text):
  if make_hashes(password) == hashed_text:
    return True
  return False


# ==========================================
# 3. GESTÃO DE BANCO DE DADOS (SQLITE)
# ==========================================
def init_db():
  conn = sqlite3.connect(DB_FILE)
  cursor = conn.cursor()

  cursor.execute("""
        CREATE TABLE IF NOT EXISTS asns (
            asn TEXT PRIMARY KEY,
            entidade TEXT,
            tipo TEXT,
            contato TEXT
        )
    """)

  cursor.execute("""
        CREATE TABLE IF NOT EXISTS blocos_ip (
            prefixo TEXT PRIMARY KEY,
            versao TEXT,
            asn_vinculado TEXT,
            finalidade TEXT,
            vlan_id TEXT,
            vrf TEXT,
            pop_localidade TEXT,
            id_circuito TEXT
        )
    """)

  cursor.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_hora TEXT,
            usuario TEXT,
            acao TEXT,
            tabela TEXT,
            detalhes TEXT
        )
    """)

  cursor.execute("""
        CREATE TABLE IF NOT EXISTS configuracoes (
            chave TEXT PRIMARY KEY,
            valor TEXT
        )
    """)

  try:
    cursor.execute("ALTER TABLE audit_logs ADD COLUMN usuario TEXT")
  except sqlite3.OperationalError:
    pass

  cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password TEXT,
            role TEXT
        )
    """)

  cursor.execute("SELECT COUNT(*) FROM users")
  if cursor.fetchone()[0] == 0:
    cursor.execute(
        "INSERT INTO users VALUES (?, ?, ?)",
        ("admin", make_hashes("admin"), "admin"),
    )

  conn.commit()
  conn.close()


def registrar_log(acao, tabela, detalhes):
  conn = sqlite3.connect(DB_FILE)
  cursor = conn.cursor()
  agora = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
  usuario_atual = st.session_state.get("username", "Sistema")
  cursor.execute(
      "INSERT INTO audit_logs (data_hora, usuario, acao, tabela, detalhes)"
      " VALUES (?, ?, ?, ?, ?)",
      (agora, usuario_atual, acao, tabela, detalhes),
  )
  conn.commit()
  conn.close()
  enviar_notificacao_webhook(
      f"🔔 *IPAM Alert* | {usuario_atual} realizou *{acao}* em *{tabela}*: {detalhes}"
  )


def autenticar_usuario(username, password):
  conn = sqlite3.connect(DB_FILE)
  cursor = conn.cursor()
  cursor.execute(
      "SELECT password, role FROM users WHERE username=?", (username.strip(),)
  )
  result = cursor.fetchone()
  conn.close()

  if result and check_hashes(password, result[0]):
    return result[1]
  return None


def carregar_dados_asn():
  conn = sqlite3.connect(DB_FILE)
  df = pd.read_sql_query("SELECT * FROM asns", conn)
  conn.close()
  df.columns = COLUNAS_ASN
  return df


def carregar_dados_ip():
  conn = sqlite3.connect(DB_FILE)
  df = pd.read_sql_query("SELECT * FROM blocos_ip", conn)
  conn.close()
  df.columns = COLUNAS_IP
  return df


def carregar_logs():
  conn = sqlite3.connect(DB_FILE)
  df = pd.read_sql_query(
      "SELECT data_hora, usuario, acao, tabela, detalhes FROM audit_logs ORDER"
      " BY id DESC",
      conn,
  )
  conn.close()
  df.columns = ["Data/Hora", "Usuário", "Ação", "Entidade", "Detalhes"]
  return df


def obter_configuracao(chave):
  conn = sqlite3.connect(DB_FILE)
  cursor = conn.cursor()
  cursor.execute("SELECT valor FROM configuracoes WHERE chave=?", (chave,))
  res = cursor.fetchone()
  conn.close()
  return res[0] if res else ""


def salvar_configuracao(chave, valor):
  conn = sqlite3.connect(DB_FILE)
  cursor = conn.cursor()
  cursor.execute(
      "INSERT OR REPLACE INTO configuracoes (chave, valor) VALUES (?, ?)",
      (chave, valor),
  )
  conn.commit()
  conn.close()


def enviar_notificacao_webhook(msg):
  webhook_url = obter_configuracao("webhook_url")
  if not webhook_url:
    return
  try:
    payload = json.dumps({"text": msg, "content": msg}).encode("utf-8")
    req = urllib.request.Request(
        webhook_url, data=payload, headers={"Content-Type": "application/json"}
    )
    urllib.request.urlopen(req, timeout=3)
  except Exception:
    pass


init_db()

# ==========================================
# 4. TELA DE LOGIN / SESSÃO
# ==========================================
if "logged_in" not in st.session_state:
  st.session_state["logged_in"] = False
if "username" not in st.session_state:
  st.session_state["username"] = ""
if "user_role" not in st.session_state:
  st.session_state["user_role"] = ""

if not st.session_state["logged_in"]:
  st.title("🔒 Acesso Restrito - IPAM Enterprise")
  st.markdown("Entre com suas credenciais para acessar o painel NOC.")

  with st.form("login_form"):
    user_input = st.text_input("Usuário")
    pass_input = st.text_input("Senha", type="password")
    submit_login = st.form_submit_button("Entrar")

    if submit_login:
      role = autenticar_usuario(user_input, pass_input)
      if role:
        st.session_state["logged_in"] = True
        st.session_state["username"] = user_input.strip()
        st.session_state["user_role"] = role
        st.success(f"Bem-vindo, {user_input}!")
        st.rerun()
      else:
        st.error("Usuário ou senha incorretos.")

  st.info("💡 **Acesso Padrão Inicial**: Usuário: `admin` | Senha: `admin`")
  st.stop()


# ==========================================
# 5. FUNÇÕES DE UTILIDADE DE REDE & FERRAMENTAS
# ==========================================
def gerar_excel_formatado():
  df_asn = carregar_dados_asn()
  df_ip = carregar_dados_ip()

  wb = openpyxl.Workbook()
  wb.remove(wb.active)

  header_fill = PatternFill(
      start_color="1F497D", end_color="1F497D", fill_type="solid"
  )
  header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
  data_font = Font(name="Segoe UI", size=10)
  center_align = Alignment(
      horizontal="center", vertical="center", wrap_text=False
  )
  left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
  thin_border = Border(
      left=Side(style="thin", color="D9D9D9"),
      right=Side(style="thin", color="D9D9D9"),
      top=Side(style="thin", color="D9D9D9"),
      bottom=Side(style="thin", color="D9D9D9"),
  )

  ws_asn = wb.create_sheet(title="ASNs")
  ws_asn.views.sheetView[0].showGridLines = True
  ws_asn.append(COLUNAS_ASN)
  for r in df_asn.itertuples(index=False):
    ws_asn.append(list(r))

  ws_ip = wb.create_sheet(title="Blocos_IP")
  ws_ip.views.sheetView[0].showGridLines = True
  ws_ip.append(COLUNAS_IP)
  for r in df_ip.itertuples(index=False):
    ws_ip.append(list(r))

  for ws in wb.worksheets:
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
      cell.fill = header_fill
      cell.font = header_font
      cell.alignment = center_align

    for col in ws.columns:
      max_len = 0
      col_letter = get_column_letter(col[0].column)
      for cell in col:
        val_str = str(cell.value) if cell.value is not None else ""
        if len(val_str) > max_len:
          max_len = len(val_str)

        if cell.row > 1:
          cell.font = data_font
          cell.border = thin_border
          ws.row_dimensions[cell.row].height = 22
          if col_letter in ["A", "B", "C", "E"]:
            cell.alignment = center_align
          else:
            cell.alignment = left_align

      ws.column_dimensions[col_letter].width = max(max_len + 6, 18)

  output = io.BytesIO()
  wb.save(output)
  return output.getvalue()


def verificar_overlap(novo_prefixo):
  try:
    nova_rede = ipaddress.ip_network(novo_prefixo.strip(), strict=False)
    df_ip = carregar_dados_ip()
    conflitos = []

    for _, row in df_ip.iterrows():
      try:
        rede_existente = ipaddress.ip_network(row["Prefixo"], strict=False)
        if (
            nova_rede.overlaps(rede_existente)
            and nova_rede.version == rede_existente.version
        ):
          conflitos.append(row["Prefixo"])
      except ValueError:
        continue
    return conflitos
  except ValueError:
    return []


def buscar_ip_global(ip_busca):
  try:
    ip_obj = ipaddress.ip_address(ip_busca.strip())
    df_ip = carregar_dados_ip()
    encontrados = []

    for _, row in df_ip.iterrows():
      try:
        rede = ipaddress.ip_network(row["Prefixo"], strict=False)
        if ip_obj in rede:
          encontrados.append(row.to_dict())
      except ValueError:
        continue
    return pd.DataFrame(encontrados)
  except ValueError:
    return pd.DataFrame()


def consultar_rdap_asn(asn_num):
  clean_asn = asn_num.upper().replace("AS", "").strip()
  url = f"https://rdap.registro.br/autnum/{clean_asn}"
  try:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=4) as response:
      if response.status == 200:
        data = json.loads(response.read().decode())
        entidade = data.get("entities", [{}])[0].get("vcardArray", [])
        nome = "N/D"
        if len(entidade) > 1:
          for item in entidade[1]:
            if item[0] == "fn":
              nome = item[3]
        return {
            "ASN": f"AS{clean_asn}",
            "Nome / Razão Social": nome,
            "Handle": data.get("handle", "N/D"),
            "Status": "Ativo / Encontrado",
        }
  except Exception:
    return {
        "ASN": f"AS{clean_asn}",
        "Mensagem": "Consulta RDAP indisponível ou ASN não localizado.",
    }


def consultar_peeringdb(asn_num):
  clean_asn = asn_num.upper().replace("AS", "").strip()
  url = f"https://www.peeringdb.com/api/net?asn={clean_asn}"
  try:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=5) as response:
      if response.status == 200:
        data = json.loads(response.read().decode())
        if data.get("data"):
          info = data["data"][0]
          return {
              "Nome PeeringDB": info.get("name"),
              "Organização": info.get("org_name"),
              "Website": info.get("website"),
              "Política de Peering": info.get("policy_general"),
              "Capacidade IPv4": info.get("info_never_via_route_servers"),
              "Local do IXP": info.get("city"),
          }
    return {"Mensagem": "ASN não localizado no PeeringDB."}
  except Exception as e:
    return {"Erro": str(e)}


def consultar_rbl(ip_str):
  rbls = ["zen.spamhaus.org", "bl.spamcop.net", "dnsbl.sorbs.net"]
  resultados = []
  try:
    ip_obj = ipaddress.ip_address(ip_str.strip())
    if ip_obj.version == 6:
      return [{"RBL": "N/A", "Status": "RBL não suporta IPv6 diretamente"}]
    rev_ip = ".".join(reversed(ip_str.split(".")))

    for rbl in rbls:
      query = f"{rev_ip}.{rbl}"
      try:
        socket.gethostbyname(query)
        resultados.append({"RBL": rbl, "Status": "⚠️ LISTADO (Blacklist)"})
      except socket.gaierror:
        resultados.append({"RBL": rbl, "Status": "✅ Limpo (OK)"})
    return resultados
  except Exception as e:
    return [{"RBL": "Erro", "Status": str(e)}]


def testar_ping(host):
  param = "-n 1" if os.name == "nt" else "-c 1"
  cmd = ["ping", param, "1", host]
  try:
    res = subprocess.run(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=2
    )
    return res.returncode == 0
  except Exception:
    return False


def testar_porta_tcp(host, port):
  try:
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(2)
    result = sock.connect_ex((host, int(port)))
    sock.close()
    return result == 0
  except Exception:
    return False


def resolver_ptr_dns(ip_str):
  try:
    return socket.gethostbyaddr(ip_str)[0]
  except Exception:
    return "Sem entrada PTR registrada"


# ==========================================
# 6. INTERFACE PRINCIPAL E NAVIGATION
# ==========================================
st.title("🌐 IPAM Enterprise - Gestão, Automação & NOC")

st.sidebar.markdown(
    f"👤 **Usuário**: `{st.session_state['username']}`"
    f" ({st.session_state['user_role']})"
)
if st.sidebar.button("🚪 Sair (Logout)"):
  st.session_state["logged_in"] = False
  st.session_state["username"] = ""
  st.session_state["user_role"] = ""
  st.rerun()

st.sidebar.divider()

df_asn = carregar_dados_asn()
df_ip = carregar_dados_ip()

opcoes_menu = [
    "📊 Dashboard & Capacidade POPs",
    "📋 Cadastrar ASN",
    "📌 Cadastrar Bloco IP",
    "✏️ Editar / Excluir Registros",
    "🧮 Calculadora & Subnetting",
    "⚡ Gerador Prefix-List BGP",
    "🛠️ Utilidades NOC (CLI, PTR & Diagnósticos)",
    "🔎 Consulta WHOIS/RDAP, PeeringDB & Blacklists",
    "📂 Importação & Exportação (CSV/Excel)",
    "📜 Histórico & Audit Log",
]

if st.session_state["user_role"] == "admin":
  opcoes_menu.append("🔐 Gestão do Sistema & Backup")

aba = st.sidebar.radio("Navegação Principal", opcoes_menu)

# ==========================================
# ABA 1: DASHBOARD & CAPACIDADE
# ==========================================
if aba == "📊 Dashboard & Capacidade POPs":
  st.header("📊 Visão Geral, Métricas e Topologia de Rede")

  col1, col2, col3, col4 = st.columns(4)
  col1.metric("ASNs Cadastrados", len(df_asn))
  col2.metric(
      "Blocos IPv4",
      len(df_ip[df_ip["Versão"] == "IPv4"]) if not df_ip.empty else 0,
  )
  col3.metric(
      "Blocos IPv6",
      len(df_ip[df_ip["Versão"] == "IPv6"]) if not df_ip.empty else 0,
  )
  col4.metric(
      "VLANs Distintas",
      df_ip["VLAN ID"].nunique() if not df_ip.empty else 0,
  )

  st.divider()

  st.subheader("🎯 Buscar um Bloco ou Host (Atalho: Ctrl + K)")
  ip_query = st.text_input(
      "Digite um IP para identificar a qual bloco pertence (ex: 192.168.1.1):",
      key="ip_search",
  )
  if ip_query:
    res = buscar_ip_global(ip_query)
    if not res.empty:
      st.success(f"O IP `{ip_query}` pertence ao bloco cadastrado abaixo:")
      st.dataframe(res, use_container_width=True)
    else:
      st.warning(f"O IP `{ip_query}` não pertence a nenhum bloco cadastrado.")

  st.divider()

  # Ocupação por POP / Localidade
  if not df_ip.empty:
    st.subheader("📍 Alocação de Blocos por POP / Localidade")
    pop_counts = df_ip["POP / Localidade"].value_counts().reset_index()
    pop_counts.columns = ["POP / Localidade", "Total de Blocos Alocados"]
    st.dataframe(pop_counts, use_container_width=True)

  # Mapa Visual de Topologia BGP
  if not df_asn.empty:
    st.subheader("🗺️ Mapa Visual de Topologia BGP & ASNs")
    dot_code = 'digraph BGP {\n  bgcolor="#0d1117";\n  node [shape=box, style="filled,rounded", fontname="Segoe UI", color="#30363d", fontcolor="#ffffff"];\n  edge [color="#58a6ff", fontname="Segoe UI", fontcolor="#8b949e"];\n'

    for _, r in df_asn.iterrows():
      asn_label = f"{r['ASN']}\\n{r['Entidade']}"
      if r["Tipo"] == "Próprio":
        dot_code += f'  "{r["ASN"]}" [label="{asn_label}", fillcolor="#1f6beb", color="#58a6ff"];\n'
      else:
        dot_code += f'  "{r["ASN"]}" [label="{asn_label}", fillcolor="#21262d"];\n'

      if r["Tipo"] != "Próprio":
        proprios = df_asn[df_asn["Tipo"] == "Próprio"]["ASN"].tolist()
        for p in proprios:
          dot_code += f'  "{r["ASN"]}" -> "{p}" [label="{r["Tipo"]}"];\n'

    dot_code += "}"
    st.graphviz_chart(dot_code)

  st.divider()

  col_t1, col_t2 = st.columns(2)
  with col_t1:
    st.subheader("📋 Tabela de ASNs")
    st.dataframe(df_asn, use_container_width=True)
  with col_t2:
    st.subheader("📌 Tabela de Blocos IP e Engenharia")
    st.dataframe(df_ip, use_container_width=True)

# ==========================================
# ABA 2: CADASTRAR ASN
# ==========================================
elif aba == "📋 Cadastrar ASN":
  st.header("📋 Registrar Novo ASN")
  with st.form("form_asn", clear_on_submit=True):
    asn_num = st.text_input("Número do ASN (ex: AS264698)")
    entidade = st.text_input("Nome da Entidade / Cliente (ex: BITCORE)")
    tipo = st.selectbox("Tipo", ["Próprio", "Cliente", "Trânsito/Upstream"])
    contato = st.text_input("E-mail / Contato NOC")

    if st.form_submit_button("Salvar ASN"):
      if asn_num and entidade:
        clean_asn = asn_num.upper().strip()
        if not clean_asn.startswith("AS"):
          clean_asn = f"AS{clean_asn}"

        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        try:
          cursor.execute(
              "INSERT INTO asns VALUES (?, ?, ?, ?)",
              (clean_asn, entidade.strip(), tipo, contato.strip()),
          )
          conn.commit()
          registrar_log("CADASTRAR", "ASNs", f"ASN {clean_asn} ({entidade})")
          st.success(f"ASN {clean_asn} salvo com sucesso!")
        except sqlite3.IntegrityError:
          st.error(f"O ASN {clean_asn} já está cadastrado no banco!")
        finally:
          conn.close()
        st.rerun()
      else:
        st.error("Preencha ASN e Entidade.")

# ==========================================
# ABA 3: CADASTRAR BLOCO IP
# ==========================================
elif aba == "📌 Cadastrar Bloco IP":
  st.header("📌 Registrar Novo Bloco IP")

  with st.form("form_ip", clear_on_submit=True):
    col_a, col_b = st.columns(2)
    with col_a:
      prefixo = st.text_input("Bloco CIDR (ex: 192.168.1.0/24)")
      asn_vinc = st.text_input("ASN Vinculado (ex: AS264698)")
      finalidade = st.text_input("Finalidade (ex: SERVIDORES / NOC)")
    with col_b:
      vlan_id = st.text_input("VLAN ID (ex: 100)")
      vrf = st.text_input("VRF (ex: VRF-INTERNET)")
      pop = st.text_input("POP / Localidade (ex: POP-SP-01)")
      id_circuito = st.text_input("ID Circuito (ex: L2VPN-CLIENTE-88412)")

    if st.form_submit_button("Salvar Bloco IP"):
      try:
        rede = ipaddress.ip_network(prefixo.strip(), strict=False)
        str_prefixo = str(rede)

        conflitos = verificar_overlap(str_prefixo)
        if conflitos:
          st.error(
              f"⚠️ Conflito de Roteamento! O bloco `{str_prefixo}` faz"
              f" sobreposição com os blocos já existentes: {conflitos}"
          )
        else:
          conn = sqlite3.connect(DB_FILE)
          cursor = conn.cursor()
          cursor.execute(
              "INSERT INTO blocos_ip VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
              (
                  str_prefixo,
                  f"IPv{rede.version}",
                  asn_vinc.upper().strip(),
                  finalidade.strip(),
                  vlan_id.strip(),
                  vrf.strip(),
                  pop.strip(),
                  id_circuito.strip(),
              ),
          )
          conn.commit()
          conn.close()
          registrar_log(
              "CADASTRAR",
              "Blocos_IP",
              f"Bloco {str_prefixo} (VLAN {vlan_id}, POP {pop})",
          )
          st.success(f"Bloco IP {str_prefixo} cadastrado com sucesso!")
          st.rerun()
      except ValueError:
        st.error("Formato CIDR inválido. Exemplo correto: 192.168.1.0/24")

# ==========================================
# ABA 4: EDITAR / EXCLUIR REGISTROS
# ==========================================
elif aba == "✏️ Editar / Excluir Registros":
  st.header("✏️ Gerenciar Registros Existentes")
  t1, t2 = st.tabs(["Gerenciar ASNs", "Gerenciar Blocos IP"])

  with t1:
    if df_asn.empty:
      st.info("Nenhum ASN cadastrado.")
    else:
      asn_sel = st.selectbox("Selecione o ASN:", df_asn["ASN"].tolist())
      row = df_asn[df_asn["ASN"] == asn_sel].iloc[0]

      with st.form("edit_asn"):
        e_entidade = st.text_input("Entidade", value=row["Entidade"])
        e_tipo = st.selectbox(
            "Tipo",
            ["Próprio", "Cliente", "Trânsito/Upstream"],
            index=["Próprio", "Cliente", "Trânsito/Upstream"].index(
                row["Tipo"]
            ),
        )
        e_contato = st.text_input("Contato", value=row["Contato"])

        c_save, c_del = st.columns(2)
        if c_save.form_submit_button("💾 Salvar Alterações"):
          conn = sqlite3.connect(DB_FILE)
          cursor = conn.cursor()
          cursor.execute(
              "UPDATE asns SET entidade=?, tipo=?, contato=? WHERE asn=?",
              (e_entidade, e_tipo, e_contato, asn_sel),
          )
          conn.commit()
          conn.close()
          registrar_log("EDITAR", "ASNs", f"ASN {asn_sel} atualizado")
          st.success(f"ASN {asn_sel} atualizado!")
          st.rerun()

        if c_del.form_submit_button("🗑️ Excluir ASN"):
          conn = sqlite3.connect(DB_FILE)
          cursor = conn.cursor()
          cursor.execute("DELETE FROM asns WHERE asn=?", (asn_sel,))
          conn.commit()
          conn.close()
          registrar_log("EXCLUIR", "ASNs", f"ASN {asn_sel} removido")
          st.warning(f"ASN {asn_sel} removido.")
          st.rerun()

  with t2:
    if df_ip.empty:
      st.info("Nenhum Bloco IP cadastrado.")
    else:
      ip_sel = st.selectbox("Selecione o Bloco IP:", df_ip["Prefixo"].tolist())
      row = df_ip[df_ip["Prefixo"] == ip_sel].iloc[0]

      with st.form("edit_ip"):
        col_e1, col_e2 = st.columns(2)
        with col_e1:
          e_asn_v = st.text_input("ASN Vinculado", value=row["ASN Vinculado"])
          e_fin = st.text_input("Finalidade", value=row["Finalidade"])
          e_vlan = st.text_input("VLAN ID", value=row["VLAN ID"])
        with col_e2:
          e_vrf = st.text_input("VRF", value=row["VRF"])
          e_pop = st.text_input(
              "POP / Localidade", value=row["POP / Localidade"]
          )
          e_circ = st.text_input("ID Circuito", value=row["ID Circuito"])

        c_s, c_d = st.columns(2)
        if c_s.form_submit_button("💾 Salvar Alterações"):
          conn = sqlite3.connect(DB_FILE)
          cursor = conn.cursor()
          cursor.execute(
              """
                        UPDATE blocos_ip 
                        SET asn_vinculado=?, finalidade=?, vlan_id=?, vrf=?, pop_localidade=?, id_circuito=?
                        WHERE prefixo=?
                    """,
              (e_asn_v, e_fin, e_vlan, e_vrf, e_pop, e_circ, ip_sel),
          )
          conn.commit()
          conn.close()
          registrar_log("EDITAR", "Blocos_IP", f"Bloco {ip_sel} atualizado")
          st.success(f"Bloco {ip_sel} atualizado!")
          st.rerun()

        if c_d.form_submit_button("🗑️ Excluir Bloco"):
          conn = sqlite3.connect(DB_FILE)
          cursor = conn.cursor()
          cursor.execute("DELETE FROM blocos_ip WHERE prefixo=?", (ip_sel,))
          conn.commit()
          conn.close()
          registrar_log("EXCLUIR", "Blocos_IP", f"Bloco {ip_sel} removido")
          st.warning(f"Bloco {ip_sel} removido.")
          st.rerun()

# ==========================================
# ABA 5: CALCULADORA & SUBNETTING
# ==========================================
elif aba == "🧮 Calculadora & Subnetting":
  st.header("🧮 Calculadora e Expansor de Sub-redes")
  pref = st.text_input("Digite o bloco/sub-rede CIDR:", value="192.168.0.0/24")

  if pref:
    try:
      rede = ipaddress.ip_network(pref.strip(), strict=False)
      c1, c2, c3, c4 = st.columns(4)
      c1.metric("Versão", f"IPv{rede.version}")
      c2.metric("Máscara CIDR", f"/{rede.prefixlen}")
      c3.metric("Total de IPs", f"{rede.num_addresses:,}")
      if rede.version == 4:
        c4.metric("Máscara Decimal", str(rede.netmask))

      st.divider()
      nova_mascara = st.slider(
          "Dividir em sub-redes menores:",
          min_value=rede.prefixlen + 1,
          max_value=32 if rede.version == 4 else 64,
          value=(
              rede.prefixlen + 2
              if rede.prefixlen < 30
              else rede.prefixlen
          ),
      )
      subredes = list(rede.subnets(new_prefix=nova_mascara))
      st.write(
          f"Bloco `{rede}` fatiado em **/{nova_mascara}** = **{len(subredes)}**"
          " sub-redes:"
      )

      dados_sub = []
      for s in subredes[:128]:
        if rede.version == 4 and s.prefixlen <= 30:
          dados_sub.append({
              "Sub-rede CIDR": str(s),
              "Network ID": str(s.network_address),
              "Primeiro IP Útil": str(s.network_address + 1),
              "Último IP Útil": str(s.broadcast_address - 1),
              "Broadcast": str(s.broadcast_address),
              "IPs Úteis": s.num_addresses - 2,
          })
        else:
          dados_sub.append({
              "Sub-rede CIDR": str(s),
              "Network ID": str(s.network_address),
              "Total de IPs": s.num_addresses,
          })

      st.dataframe(pd.DataFrame(dados_sub), use_container_width=True)
    except ValueError:
      st.error("Bloco CIDR inválido.")

# ==========================================
# ABA 6: GERADOR PREFIX-LIST BGP
# ==========================================
elif aba == "⚡ Gerador Prefix-List BGP":
  st.header("⚡ Gerador de Prefix-Lists BGP")
  if df_ip.empty:
    st.info("Nenhum bloco cadastrado.")
  else:
    vendor = st.selectbox(
        "Vendor do Roteador:",
        [
            "Cisco (IOS/IOS-XE)",
            "Huawei (VRP)",
            "Juniper (Junos)",
            "MikroTik RouterOS v7",
        ],
    )
    nome_pl = st.text_input("Nome da Prefix-List:", "PL-BGP-ANNOUNCE")
    blocos_sel = st.multiselect(
        "Blocos para Anúncio:",
        options=df_ip["Prefixo"].tolist(),
        default=df_ip["Prefixo"].tolist(),
    )

    if blocos_sel and nome_pl:
      script_out = ""
      for idx, p in enumerate(blocos_sel, start=10):
        r = ipaddress.ip_network(p, strict=False)
        if vendor == "Cisco (IOS/IOS-XE)":
          script_out += f"{'ip' if r.version == 4 else 'ipv6'} prefix-list {nome_pl} seq {idx} permit {p}\n"
        elif vendor == "Huawei (VRP)":
          script_out += f"{'ip' if r.version == 4 else 'ipv6'} ip-prefix {nome_pl} index {idx} permit {r.network_address} {r.prefixlen}\n"
        elif vendor == "Juniper (Junos)":
          script_out += f"set policy-options prefix-list {nome_pl} {p}\n"
        elif vendor == "MikroTik RouterOS v7":
          script_out += f'/routing filter rule add chain={nome_pl} rule="if (dst in {p}) {{ accept }}"\n'

      st.code(script_out, language="text")
      botao_copiar(script_out, id_unico="btnCopyPL")

# ==========================================
# ABA 7: UTILIDADES NOC & DIAGNÓSTICOS
# ==========================================
elif aba == "🛠️ Utilidades NOC (CLI, PTR & Diagnósticos)":
  st.header("🛠️ Automação e Ferramentas NOC")
  tab_cli, tab_ptr, tab_diag = st.tabs([
      "Gerador CLI Interfaces",
      "Gerador DNS Reverso (PTR)",
      "⚡ Diagnósticos Ativos (Ping / Port / DNS)",
  ])

  with tab_cli:
    if df_ip.empty:
      st.info("Cadastre blocos IP para utilizar o gerador.")
    else:
      bloco_cli = st.selectbox(
          "Selecione o Bloco IP:", df_ip["Prefixo"].tolist()
      )
      row_cli = df_ip[df_ip["Prefixo"] == bloco_cli].iloc[0]
      vendor_cli = st.selectbox(
          "Selecione o Vendor:",
          [
              "Cisco (IOS-XE)",
              "Huawei (VRP)",
              "Juniper (Junos)",
              "MikroTik RouterOS v7",
          ],
      )

      r = ipaddress.ip_network(bloco_cli, strict=False)
      ip_gw = r.network_address + 1
      vlan = row_cli["VLAN ID"] or "100"
      desc = f"{row_cli['Finalidade']} - {row_cli['ID Circuito']}"

      st.subheader("Script de Configuração de Interface:")
      cli_code = ""

      if vendor_cli == "Cisco (IOS-XE)":
        cli_code = f"""interface GigabitEthernet0/0/0.{vlan}
 encapsulation dot1Q {vlan}
 description {desc}
 ip address {ip_gw} {r.netmask}
 no shutdown"""
      elif vendor_cli == "Huawei (VRP)":
        cli_code = f"""interface GigabitEthernet0/1/0.{vlan}
 dot1q termination vid {vlan}
 description {desc}
 ip address {ip_gw} {r.prefixlen}
 arp broadcast enable"""
      elif vendor_cli == "Juniper (Junos)":
        cli_code = f"""set interfaces ge-0/0/0 unit {vlan} vlan-id {vlan}
set interfaces ge-0/0/0 unit {vlan} description "{desc}"
set interfaces ge-0/0/0 unit {vlan} family inet address {ip_gw}/{r.prefixlen}"""
      elif vendor_cli == "MikroTik RouterOS v7":
        cli_code = f"""/interface vlan add name=vlan{vlan} vlan-id={vlan} interface=ether1 comment="{desc}"
/ip address add address={ip_gw}/{r.prefixlen} interface=vlan{vlan}"""

      st.code(cli_code, language="text")
      botao_copiar(cli_code, id_unico="btnCopyCLI")

  with tab_ptr:
    if df_ip.empty:
      st.info("Cadastre blocos IP para utilizar a ferramenta.")
    else:
      bloco_ptr = st.selectbox(
          "Selecione o Bloco para Zona Reverso:",
          df_ip["Prefixo"].tolist(),
          key="ptr_box",
      )
      dominio_ptr = st.text_input("Domínio de Host (ex: noc.provedor.com.br)")

      if dominio_ptr:
        r_ptr = ipaddress.ip_network(bloco_ptr, strict=False)
        st.subheader("Entradas de Zona BIND9 / DNS Reverso:")
        ptr_out = ""

        if r_ptr.version == 4 and r_ptr.prefixlen <= 30:
          for ip in list(r_ptr.hosts())[:32]:
            octetos = str(ip).split(".")
            ptr_out += (
                f"{octetos[3]}\tIN\tPTR\thost-{octetos[3]}.{dominio_ptr}.\n"
            )
        else:
          ptr_out = (
              f"; Zona IPv6/Sub-rede personalizada:\n; Prefixo:"
              f" {r_ptr.network_address} PTR {dominio_ptr}"
          )

        st.code(ptr_out, language="text")
        botao_copiar(ptr_out, id_unico="btnCopyPTR")

  with tab_diag:
    st.subheader("📡 Diagnósticos Ativos em Tempo Real")
    col_d1, col_d2, col_d3 = st.columns(3)

    with col_d1:
      st.markdown("#### Teste de ICMP Ping")
      host_ping = st.text_input("IP para Ping (ex: 8.8.8.8)")
      if st.button("Disparar Ping"):
        if host_ping:
          res_ping = testar_ping(host_ping)
          if res_ping:
            st.success(f" Host `{host_ping}` respondendo!")
          else:
            st.error(f"❌ Host `{host_ping}` não respondeu.")

    with col_d2:
      st.markdown("#### Teste de Porta TCP")
      host_tcp = st.text_input("IP / Host TCP")
      porta_tcp = st.number_input("Porta (ex: 22, 80, 443)", value=22, step=1)
      if st.button("Testar Porta TCP"):
        if host_tcp:
          res_tcp = testar_porta_tcp(host_tcp, porta_tcp)
          if res_tcp:
            st.success(f" Porta `{porta_tcp}` aberta em `{host_tcp}`!")
          else:
            st.error(f"❌ Porta `{porta_tcp}` fechada/timeout em `{host_tcp}`.")

    with col_d3:
      st.markdown("#### Checagem PTR DNS Reverso")
      ip_ptr_test = st.text_input("IP para consulta PTR")
      if st.button("Resolver PTR"):
        if ip_ptr_test:
          res_ptr = resolver_ptr_dns(ip_ptr_test)
          st.info(f" Hostname PTR: `{res_ptr}`")

# ==========================================
# ABA 8: CONSULTAS BGP, RDAP & BLACKLISTS
# ==========================================
elif aba == "🔎 Consulta WHOIS/RDAP, PeeringDB & Blacklists":
  st.header("🔎 Inteligência BGP, RDAP e Checagem de Blacklists")

  t_rdap, t_pdb, t_rbl = st.tabs(
      ["Consulta RDAP/Registro.br", "PeeringDB Lookups", "Checagem de RBL"]
  )

  with t_rdap:
    asn_query = st.text_input(
        "Digite o número do ASN para consulta (ex: 264698)"
    )
    if st.button("Consultar no Registro.br / RDAP"):
      if asn_query:
        dados_rdap = consultar_rdap_asn(asn_query)
        st.json(dados_rdap)

  with t_pdb:
    asn_pdb = st.text_input("Digite o ASN para consultar no PeeringDB:")
    if st.button("Buscar PeeringDB"):
      if asn_pdb:
        dados_pdb = consultar_peeringdb(asn_pdb)
        st.json(dados_pdb)

  with t_rbl:
    ip_rbl = st.text_input("Digite o IPv4 para verificar em Blacklists (RBL):")
    if st.button("Verificar Blacklists"):
      if ip_rbl:
        res_rbl = consultar_rbl(ip_rbl)
        st.dataframe(pd.DataFrame(res_rbl), use_container_width=True)

# ==========================================
# ABA 9: IMPORTAÇÃO & EXPORTAÇÃO
# ==========================================
elif aba == "📂 Importação & Exportação (CSV/Excel)":
  st.header("📂 Gerenciamento de Dados em Massa")

  st.subheader("📥 Exportação de Planilha Excel Estilizada")
  excel_data = gerar_excel_formatado()
  st.download_button(
      label="📥 Baixar IPAM_Database.xlsx",
      data=excel_data,
      file_name="ipam_database.xlsx",
      mime=(
          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
      ),
  )

  st.divider()

  st.subheader("📤 Importar Blocos IP via arquivo CSV")
  st.markdown(
      "O arquivo CSV deve conter o cabeçalho exatamente assim: `Prefixo, Versão,"
      " ASN Vinculado, Finalidade, VLAN ID, VRF, POP / Localidade, ID Circuito`"
  )
  uploaded_file = st.file_uploader("Escolha um arquivo CSV", type=["csv"])

  if uploaded_file is not None:
    try:
      df_upload = pd.read_csv(uploaded_file)
      st.write("Prévia dos dados:")
      st.dataframe(df_upload.head(), use_container_width=True)

      if st.button("Confirmar Importação de CSV"):
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        qtd = 0
        for _, row in df_upload.iterrows():
          try:
            cursor.execute(
                "INSERT INTO blocos_ip VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    str(row["Prefixo"]).strip(),
                    str(row["Versão"]).strip(),
                    str(row["ASN Vinculado"]).strip(),
                    str(row["Finalidade"]).strip(),
                    str(row["VLAN ID"]).strip(),
                    str(row["VRF"]).strip(),
                    str(row["POP / Localidade"]).strip(),
                    str(row["ID Circuito"]).strip(),
                ),
            )
            qtd += 1
          except Exception:
            continue
        conn.commit()
        conn.close()
        registrar_log(
            "IMPORTAR", "Blocos_IP", f"Importados {qtd} blocos via CSV"
        )
        st.success(f"{qtd} blocos importados com sucesso!")
        st.rerun()
    except Exception as e:
      st.error(f"Erro ao processar arquivo CSV: {e}")

# ==========================================
# ABA 10: HISTÓRICO & AUDIT LOG
# ==========================================
elif aba == "📜 Histórico & Audit Log":
  st.header("📜 Histórico de Alterações (Audit Log)")
  df_logs = carregar_logs()
  if df_logs.empty:
    st.info("Nenhum registro de log até o momento.")
  else:
    st.dataframe(df_logs, use_container_width=True)

# ==========================================
# ABA 11: GESTÃO DO SISTEMA & BACKUP (ADMIN ONLY)
# ==========================================
elif aba == "🔐 Gestão do Sistema & Backup":
  if st.session_state["user_role"] != "admin":
    st.error("⚠️ Acesso Negado! Apenas administradores podem acessar esta área.")
    st.stop()

  st.header("🔐 Configurações do Sistema e Backup")

  tab_backup, tab_notif, tab_users = st.tabs([
      "💾 Backup / Restauração DB",
      "🔔 Configurar Notificações Webhook",
      "👥 Controle de Usuários",
  ])

  with tab_backup:
    st.subheader("Download do Banco de Dados SQLite")
    if os.path.exists(DB_FILE):
      with open(DB_FILE, "rb") as f:
        st.download_button(
            label="📥 Baixar Backup `ipam_database.db`",
            data=f,
            file_name=f"ipam_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
            mime="application/x-sqlite3",
        )

  with tab_notif:
    st.subheader("Integração com Webhook (Telegram / Slack / Discord / N8N)")
    webhook_atual = obter_configuracao("webhook_url")
    novo_webhook = st.text_input(
        "URL do Webhook para alertas:", value=webhook_atual
    )
    if st.button("Salvar Webhook"):
      salvar_configuracao("webhook_url", novo_webhook.strip())
      st.success("URL de Webhook salva com sucesso!")

  with tab_users:
    t_list, t_add, t_edit = st.tabs(
        ["📋 Usuários Cadastrados", "➕ Novo Usuário", "✏️ Editar Usuário"]
    )

    with t_list:
      conn = sqlite3.connect(DB_FILE)
      df_users = pd.read_sql_query("SELECT username, role FROM users", conn)
      conn.close()
      df_users.columns = ["Usuário (Login)", "Perfil de Acesso (Role)"]
      st.dataframe(df_users, use_container_width=True)

    with t_add:
      with st.form("form_novo_usuario", clear_on_submit=True):
        new_user = st.text_input("Novo Usuário (Login)")
        new_pass = st.text_input("Senha", type="password")
        new_role = st.selectbox("Perfil", ["operator", "admin"])

        if st.form_submit_button("Cadastrar Usuário"):
          if new_user and new_pass:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            try:
              cursor.execute(
                  "INSERT INTO users VALUES (?, ?, ?)",
                  (new_user.strip(), make_hashes(new_pass), new_role),
              )
              conn.commit()
              registrar_log(
                  "CADASTRAR",
                  "Usuários",
                  f"Usuário {new_user.strip()} cadastrado como {new_role}",
              )
              st.success(
                  f"Usuário `{new_user.strip()}` registrado com sucesso!"
              )
              st.rerun()
            except sqlite3.IntegrityError:
              st.error("Este nome de usuário já existe no sistema.")
            finally:
              conn.close()

    with t_edit:
      conn = sqlite3.connect(DB_FILE)
      df_u_edit = pd.read_sql_query("SELECT username, role FROM users", conn)
      conn.close()

      if not df_u_edit.empty:
        lista_usuarios = df_u_edit["username"].tolist()
        user_selecionado = st.selectbox("Selecione o Usuário:", lista_usuarios)
        role_atual = df_u_edit[df_u_edit["username"] == user_selecionado][
            "role"
        ].values[0]

        with st.form("form_editar_usuario"):
          novo_perfil = st.selectbox(
              "Perfil de Acesso",
              ["operator", "admin"],
              index=0 if role_atual == "operator" else 1,
          )
          nova_senha = st.text_input(
              "Nova Senha (deixe em branco se não quiser alterar)",
              type="password",
          )

          btn_salvar, btn_excluir = st.columns(2)

          if btn_salvar.form_submit_button("💾 Salvar Alterações"):
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            if nova_senha.strip():
              cursor.execute(
                  "UPDATE users SET role=?, password=? WHERE username=?",
                  (novo_perfil, make_hashes(nova_senha.strip()), user_selecionado),
              )
            else:
              cursor.execute(
                  "UPDATE users SET role=? WHERE username=?",
                  (novo_perfil, user_selecionado),
              )
            conn.commit()
            conn.close()
            registrar_log(
                "EDITAR",
                "Usuários",
                f"Perfil do usuário {user_selecionado} atualizado",
            )
            st.success("Usuário atualizado com sucesso!")
            st.rerun()

          if btn_excluir.form_submit_button("🗑️ Excluir Usuário"):
            if user_selecionado == st.session_state["username"]:
              st.error("Você não pode excluir sua própria conta logada.")
            else:
              conn = sqlite3.connect(DB_FILE)
              cursor = conn.cursor()
              cursor.execute(
                  "DELETE FROM users WHERE username=?", (user_selecionado,)
              )
              conn.commit()
              conn.close()
              registrar_log(
                  "EXCLUIR", "Usuários", f"Usuário {user_selecionado} excluído"
              )
              st.warning("Usuário removido!")
              st.rerun()